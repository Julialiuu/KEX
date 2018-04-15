import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
import os,csv

not_available = "N/A"

def get_links_from_bom_copypaste():
    import openpyxl as excel_loader

    #bom_copypaste_file = os.path.join("..", "..", "..", "data", "bom_copypaste.xlsm")
    work_book = excel_loader.load_workbook("bom_copypaste.xlsm")
    work_sheet = work_book.get_sheet_by_name("BOM")

    c = 3
    row = 6
    link_list = []
    for r in range(row, 730):
        cell_val = work_sheet.cell(row=r, column=c).value
        try:
            cell_link = work_sheet.cell(row=r, column=c).hyperlink.target
        except AttributeError:
            cell_link = not_available
        new_line = [cell_val, cell_link]
        link_list.append(new_line)
    logger.info("international link retrieval Done")
    return link_list

def write_to_file(file_name, row_list, header=None):
    import unicodedata
    with open(file_name, "w", newline='\n') as out_file:
        csv_writer = csv.writer(out_file, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        if header != None:
            csv_writer.writerow(header)
        for row in row_list:
            try:
                csv_writer.writerow(row)
            except UnicodeEncodeError:
                #This exception should only happen when writing cell_val from excel bom_copypaste
                cell_val = row[0]
                cell_val = unicodedata.normalize("NFKD", cell_val)
                row[0] = cell_val
                csv_writer.writerow(row)

#Utility functions for cleaning data
def _clean_mpaa(mpaa):
    if mpaa == "Unrated":
        return not_available
    else:
        return mpaa

def _clean_prod_budget(budget):
    if budget == not_available:
        return not_available
    else:
        #Remove $
        budget = budget[1:]
        budget = budget.replace(",", "")
        #Convert million ot number
        if budget[-8:] == " million":
            budget = budget.split()
            budget = float(budget[0]) * 1000000
            budget = int(budget)
        return budget

def _clean_gross(number):
    if number != not_available:
        # remove $ at the beginning
        number = number[1:]
        # Replace
        number = number.replace(",", "")
        return number
    else:
        return not_available


'''
 manual_modification = {"02:22:00":"2:22",
                           "2018-09-11 00:00:00":"September 11"}
#Manual modify film value as openpyxl changes the format
        if cell_val in manual_modification.keys():
            cell_val = manual_modification[cell_val]
            print(cell_val, work_sheet.cell(row=r, column=c).hyperlink.target)
'''